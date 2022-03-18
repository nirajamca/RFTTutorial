# **********************************************************************************
# Function to read values from data spreadsheet
# Name:         fncGetValues
# Attributes:   loc - location of the datafile
#               sheetname - sheet name
#               testcase - test case name
# Returns:      a dictionary with all the data element matching with testcase name
# Usage:        return fncGetValues(loc, sheetName, testCase)
# *********************************************************************************
def fncGetValues(loc, sheetName, testCase):
    # import openpyxl module to use excel operations
    import openpyxl
    from openpyxl import load_workbook

    # Create workbook object and capture the active sheet
    wb = load_workbook(loc, data_only=True)
    sheet = wb.active

    # Designate the actual sheetname
    sheet = wb[sheetName]

    # Leaving the header row, go through each line to find a match with test case you want to execute
    for i in range(2, sheet.max_row + 1):
        if (sheet.cell(row=i, column=1).value == testCase):

            # if the matching test case is found, create a dictionary with data in second column
            getValues = {sheet.cell(row=1, column=2).value: sheet.cell(row=i, column=2).value}

            # Once the data dictionary is added, keep on adding rest of the items in that row till the end of columns
            for j in range(3, sheet.max_column + 1):
                getValues.update({sheet.cell(row=1, column=j).value: sheet.cell(row=i, column=j).value})

            break

    # Return the dictionary
    return getValues

# **********************************************************************************
# Function to return data based on spreadsheet and test case name
# Name:         fncGetData
# Attributes:   testcase - test case name
# Returns:      a dictionary with all the data element matching with testcase name
# Usage:        myData = fncGetData(testcase)
# *********************************************************************************
def fncGetData(testcase):
    import platform

    # To make it work on both Windows and Mac, call module system
    if platform.system() == 'Windows':
        loc = 'Data\\OrangeHRMData.xlsx'
    elif platform.system() == 'Darwin':
        loc = 'Data/OrangeHRMData.xlsx'

    # Default sheet name
    SheetName = 'OrangeHRM'

    # Return the dictionary
    return fncGetValues(loc, SheetName, testcase)

